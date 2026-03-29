import pdfplumber
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import os
from datetime import datetime

# Expresiones regulares para extraer los datos
# Patrón: Código (números) - Descripción - Cantidad (números)
PATRON_LINEA:str = r'^\s*(\d+)\s+(.+?)\s+(\d+)\s*$'
fecha_hoy:str = datetime.now().strftime("%d%m%Y")
nueva_carpeta:str = os.path.join("..", "Descuentos", f"Descuento-{fecha_hoy}")

def borrarPantalla():
    """
    Limpia la pantalla de la consola
    """
    os.system('cls' if os.name == 'nt' else 'clear')

def crear_carpeta_salida(nueva_carpeta):
    """
    Crea una carpeta de salida para los archivos Excel
    """
    if not os.path.exists("../Descuentos"):
        os.makedirs("../Descuentos")
        print(f"📁 Carpeta creada: Descuentos ✅")
    else:
        print(f"📁 Carpeta ya existe: Descuentos ")

    if not os.path.exists(nueva_carpeta):
        os.makedirs(nueva_carpeta)
        print(f"📁 Carpeta creada: {nueva_carpeta} ✅")
    else:
        print(f"📁 Carpeta ya existe: {nueva_carpeta}")




def extraer_productos_pdf(ruta_pdf):
    """
    Extrae productos de un PDF siguiendo el patrón:
    Código | Descripción | Cantidad
    """
    productos: list = []
    
    try:
        with pdfplumber.open(ruta_pdf) as pdf:
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text()
                
                # Dividir por líneas
                lineas = text.split('\n')
                
                # Buscar líneas con el patrón de producto
                # El patrón es: número at inicio, descripción, marca, cantidad al final
                for linea in lineas:
                    linea = linea.strip()
                    
                    # Intentar coincidencia simple mejorada
                    # Patrón: inicia con números, finalmente números (cantidad)
                    if re.match(r'^\d+\s', linea) and re.search(r'\s+\d+\s*$', linea):
                        partes = linea.split()
                        
                        if len(partes) >= 3:
                            codigo = partes[0]
                            cantidad_str = partes[-1]
                            
                            # Validar que cantidad sea número
                            try:
                                cantidad = int(cantidad_str)
                                # Descripción es todo lo intermedio
                                descripcion = ' '.join(partes[1:-1])
                                
                                productos.append({
                                    'codigo': codigo,
                                    'descripcion': descripcion,
                                    'cantidad': cantidad,
                                    'pdf': Path(ruta_pdf).name
                                })
                            except ValueError:
                                pass
    except Exception as e:
        print(f"Error al procesar {ruta_pdf}: {e}")
    
    return productos

def consolidar_productos(lista_productos):
    """
    Agrupa productos por código y suma cantidades
    """
    consolidado = {}
    
    for prod in lista_productos:
        codigo = prod['codigo']
        
        if codigo not in consolidado:
            consolidado[codigo] = {
                'codigo': codigo,
                'descripcion': prod['descripcion'],
                'cantidad': prod['cantidad'],
                'pdfs': [prod['pdf']]
            }
        else:
            consolidado[codigo]['cantidad'] += prod['cantidad']
            if prod['pdf'] not in consolidado[codigo]['pdfs']:
                consolidado[codigo]['pdfs'].append(prod['pdf'])
    
    return list(consolidado.values())

def exportar_excel(productos_consolidados, archivo_salida):
    """
    Exporta los productos consolidados a un archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Descuento-{fecha_hoy}"
    
    # Encabezados
    encabezados = ['Código', 'Descripción', 'Cantidad', 'PDFs']
    ws.append(encabezados)
    
    # Estilo encabezado
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Datos
    for prod in productos_consolidados:
        ws.append([
            prod['codigo'],
            prod['descripcion'],
            prod['cantidad'],
            ', '.join(prod['pdfs'])
        ])
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 12
    
    # Calcular ancho dinámico para columna PDFs
    max_pdf_width = len("PDFs")  # Ancho mínimo del encabezado
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            if cell.value:
                # Calcular ancho basado en el contenido + margen
                contenido_width = len(str(cell.value)) / 2 + 2
                max_pdf_width = max(max_pdf_width, contenido_width)
    
    ws.column_dimensions['D'].width = max_pdf_width
    
    # Alineación
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].alignment = Alignment(horizontal="center")  # Código
        row[1].alignment = Alignment(horizontal="left", wrap_text=True)  # Descripción
        row[2].alignment = Alignment(horizontal="center")  # Cantidad
        row[3].alignment = Alignment(horizontal="left", wrap_text=True)  # PDFs
    
    # Aplicar bordes a todas las celdas utilizadas
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
    
    wb.save(archivo_salida)
    print(f"✅ Excel guardado: {archivo_salida}")

# EJECUTAR
borrarPantalla()
print("=" * 70)
print("EXTRACTOR DE PRODUCTOS - REMITOS PDF")
print("=" * 70)

# Crear carpeta de salida
crear_carpeta_salida(nueva_carpeta)

# Buscar todos los PDFs
pdfs = list(Path('.').glob('*.pdf'))
print(f"\n📄 PDFs encontrados: {len(pdfs)}")

todos_productos = []

for pdf in pdfs:
    print(f"  Procesando: {pdf.name}")
    productos = extraer_productos_pdf(str(pdf))
    todos_productos.extend(productos)
    print(f"    ✓ {len(productos)} productos extraídos")

print(f"\n📊 Total de líneas de productos: {len(todos_productos)}")

# Consolidar duplicados
productos_consolidados = consolidar_productos(todos_productos)
print(f"📊 Productos únicos (consolidados): {len(productos_consolidados)}")

# Mostrar preview
print("\n--- VISTA PREVIA ---")
for prod in productos_consolidados[:10]:
    print(f"  {prod['codigo']}: {prod['descripcion'][:40]} - Cantidad: {prod['cantidad']}")

# Exportar a Excel
fecha_hoy = datetime.now().strftime("%d%m%Y")
archivo_excel = os.path.join(nueva_carpeta, f"Descuento-{fecha_hoy}.xlsx")
exportar_excel(productos_consolidados, archivo_excel)

print(f"\n✅ Proceso completado!")
print(f"   Total de productos únicos: {len(productos_consolidados)}")
print(f"   Cantidad total: {sum(p['cantidad'] for p in productos_consolidados)}")



input("\nPresiona Enter para salir...")
